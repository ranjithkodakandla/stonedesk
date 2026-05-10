from copy import deepcopy
from datetime import date, datetime
from io import BytesIO
import os
import time
from typing import Any, Dict, List, Optional, Tuple

import certifi
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Response, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pymongo import MongoClient, ReturnDocument
from pydantic import BaseModel
from .pdf_parser import parse_pdf
from .services.container_planner import build_container_plan
from .services.deterministic_packing import (
    CATEGORY_CONFIG as DET_CATEGORY_CONFIG,
    _piece_weight as det_piece_weight,
    build_packing_families,
    discover_dispatch_hierarchy,
)
from .services.planner_v3 import enrich_layout_with_crates, run_v3_planner
from .services.planner_v3.container_layout import linear_manual_sort_placements
from .services.planning_engine import (
    COLOR_DENSITIES,
    build_planning_snapshot,
    piece_destination_key,
    piece_weight as planning_piece_weight,
    weight_factor as planning_weight_factor,
)

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

MONGODB_URI = os.getenv("MONGODB_URI", "mongodb+srv://virgin_db_user:iddh38iXtoKpt1We@cluster0.t9reftj.mongodb.net/?appName=Cluster0")
MONGODB_DB = os.getenv("MONGODB_DB", "virgin")
ALLOW_MEMORY_FALLBACK = os.getenv("ALLOW_MEMORY_FALLBACK", "false").lower() in {"1", "true", "yes"}

class InMemoryCollection:
    def __init__(self):
        self.docs = []
        self.indexes = []

    def create_index(self, *args, **kwargs):
        self.indexes.append((args, kwargs))

    def _matches(self, doc, query):
        for key, value in (query or {}).items():
            if isinstance(value, dict) and "$in" in value:
                if doc.get(key) not in value["$in"]:
                    return False
            elif doc.get(key) != value:
                return False
        return True

    def _project(self, doc, projection):
        if not projection:
            return deepcopy(doc)
        if isinstance(projection, dict):
            include = {k for k, v in projection.items() if v and k != "_id"}
            if include:
                return {k: deepcopy(doc[k]) for k in include if k in doc}
            return deepcopy(doc)
        return deepcopy(doc)

    def insert_one(self, doc):
        self.docs.append(deepcopy(doc))
        return doc

    def insert_many(self, docs):
        for doc in docs:
            self.insert_one(doc)

    def find(self, query=None, projection=None):
        return [self._project(doc, projection) for doc in self.docs if self._matches(doc, query)]

    def find_one(self, query=None, projection=None, sort=None):
        docs = self.find(query, projection)
        if sort:
            key, direction = sort[0]
            docs = sorted(docs, key=lambda d: d.get(key), reverse=direction < 0)
        return docs[0] if docs else None

    def update_one(self, query, update):
        for doc in self.docs:
            if self._matches(doc, query):
                if "$set" in update:
                    doc.update(deepcopy(update["$set"]))
                return

    def delete_one(self, query):
        for idx, doc in enumerate(self.docs):
            if self._matches(doc, query):
                self.docs.pop(idx)
                return

    def delete_many(self, query):
        self.docs = [doc for doc in self.docs if not self._matches(doc, query)]

    def find_one_and_update(self, query, update, upsert=False, return_document=None):
        doc = self.find_one(query)
        if doc is None and upsert:
            doc = {"_id": query.get("_id")}
            self.docs.append(doc)
        if doc is None:
            return None
        target = next((item for item in self.docs if self._matches(item, query)), None)
        if target is None:
            return None
        if "$inc" in update:
            for key, value in update["$inc"].items():
                target[key] = target.get(key, 0) + value
        return deepcopy(target)


def build_store():
    try:
        if not MONGODB_URI:
            raise ValueError("MONGODB_URI environment variable is missing.")

        connect_kwargs = {"serverSelectionTimeoutMS": 1500}
        if MONGODB_URI.startswith("mongodb+srv"):
            connect_kwargs["tlsCAFile"] = certifi.where()
        mongo_client = MongoClient(MONGODB_URI, **connect_kwargs)
        mongo_client.admin.command("ping")
        print(f"Connected to Production MongoDB: {MONGODB_URI.split('@')[-1] if '@' in MONGODB_URI else 'Remote Host'}")
        mongo_db = mongo_client[MONGODB_DB]
        store = {
            "projects": mongo_db["projects"],
            "pieces": mongo_db["pieces"],
            "crates": mongo_db["crates"],
            "assignments": mongo_db["assignments"],
            "counters": mongo_db["counters"],
            "upload_drafts": mongo_db["upload_drafts"],
        }
        return store, "mongo"
    except Exception as e:
        print(f"Production Database Connection Failed: {e}")
        if not ALLOW_MEMORY_FALLBACK:
            raise
        return {
            "projects": InMemoryCollection(),
            "pieces": InMemoryCollection(),
            "crates": InMemoryCollection(),
            "assignments": InMemoryCollection(),
            "counters": InMemoryCollection(),
            "upload_drafts": InMemoryCollection(),
        }, "memory"


store, STORE_MODE = build_store()
projects_col = store["projects"]
pieces_col = store["pieces"]
crates_col = store["crates"]
assignments_col = store["assignments"]
counters_col = store["counters"]
upload_drafts_col = store["upload_drafts"]


def ensure_indexes() -> None:
    if not hasattr(projects_col, "create_index"):
        return

    projects_col.create_index("id")
    pieces_col.create_index("id")
    pieces_col.create_index([("project_id", 1), ("id", 1)])
    crates_col.create_index("id")
    crates_col.create_index([("project_id", 1), ("id", 1)])
    assignments_col.create_index("id")
    assignments_col.create_index([("project_id", 1), ("piece_id", 1)])
    assignments_col.create_index([("project_id", 1), ("crate_id", 1)])


ensure_indexes()


def next_sequence(name: str) -> int:
    counter = counters_col.find_one_and_update(
        {"_id": name},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return int(counter["seq"])


def utc_now() -> datetime:
    return datetime.utcnow()


def as_iso(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def project_response(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not doc:
        return None
    return {
        "id": doc["id"],
        "name": doc.get("name", ""),
        "customer": doc.get("customer", ""),
        "job_number": doc.get("job_number", ""),
        "material": doc.get("material", "Granite"),
        "thickness": doc.get("thickness", "3CM"),
        "stone_color": doc.get("stone_color", ""),
        "crate_wood_type": doc.get("crate_wood_type", "Pine"),
        "crate_wood_thickness": doc.get("crate_wood_thickness", 1.25),
        "preferred_container_mode": doc.get("preferred_container_mode", "recommended"),
        "date": doc.get("date", date.today().isoformat()),
        "flat_format": doc.get("flat_format", "3-digit"),
        "description_thickness_map": doc.get("description_thickness_map", {}),
        "status": doc.get("status", "draft"),
        "dispatch_selection": doc.get("dispatch_selection", {}),
        "planner_v3_layout": doc.get("planner_v3_layout"),
        "planner_v3_containers": doc.get("planner_v3_containers") or [],
        "planner_v3_summary": doc.get("planner_v3_summary"),
        "planner_v3_container_optimization": doc.get("planner_v3_container_optimization"),
        "delivery_payload_cap_kg": float(doc.get("delivery_payload_cap_kg") or 24000),
        "created_at": as_iso(doc.get("created_at")),
        "updated_at": as_iso(doc.get("updated_at")),
    }


def piece_response(doc: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": doc["id"],
        "project_id": doc["project_id"],
        "part": doc.get("part", ""),
        "part_no": doc.get("part_no", ""),
        "category": doc.get("category", ""),
        "drawing": doc.get("drawing", ""),
        "length": doc.get("length", 0.0),
        "width": doc.get("width", 0.0),
        "thickness": doc.get("thickness", "3CM"),
        "qty": doc.get("qty", 1),
        "unit": doc.get("unit", ""),
        "edge": doc.get("edge", ""),
        "edge_area": doc.get("edge_area", ""),
        "edge_polish_machine": doc.get("edge_polish_machine", 0.0),
        "edge_map": doc.get("edge_map", {}),
        "edge_polish_manual": doc.get("edge_polish_manual", ""),
        "radius": doc.get("radius", "-"),
        "radius_value": doc.get("radius_value", ""),
        "radius_corners": doc.get("radius_corners", {}),
        "shape_type": doc.get("shape_type", ""),
        "sink_type": doc.get("sink_type", "No Sink"),
        "sink_cut": doc.get("sink_cut", "-"),
        "tap_holes": doc.get("tap_holes", "-"),
        "grooves": doc.get("grooves", "-"),
        "fragility": doc.get("fragility", "Standard"),
        "orientation": doc.get("orientation", "Auto"),
        "delivery_priority": doc.get("delivery_priority", "Standard"),
        "stack_preference": doc.get("stack_preference", "Auto"),
        "weight_override": doc.get("weight_override", 0.0),
        "notes": doc.get("notes", ""),
        "building": doc.get("building", ""),
        "floor": doc.get("floor", ""),
        "flat": doc.get("flat", ""),
        "created_at": as_iso(doc.get("created_at")),
    }


def crate_response(doc: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": doc["id"],
        "project_id": doc["project_id"],
        "crate_id": doc.get("crate_id", ""),
        "name": doc.get("name", ""),
        "max_weight": doc.get("max_weight", 1000.0),
        "internal_length": doc.get("internal_length", 0.0),
        "internal_width": doc.get("internal_width", 0.0),
        "internal_height": doc.get("internal_height", 0.0),
        "external_length": doc.get("external_length", 0.0),
        "external_width": doc.get("external_width", 0.0),
        "external_height": doc.get("external_height", 0.0),
        "sqft": doc.get("sqft", 0.0),
        "weight": doc.get("weight", 0.0),
        "dispatch_order": doc.get("dispatch_order", 0),
        "packing_mode": doc.get("packing_mode", ""),
        "packing_family": doc.get("packing_family", ""),
        "splash_layer": doc.get("splash_layer", False),
        "main_layer_piece_ids": doc.get("main_layer_piece_ids", []),
        "splash_layer_piece_ids": doc.get("splash_layer_piece_ids", []),
        "packing_warnings": doc.get("packing_warnings", []),
        "grouping_reason": doc.get("grouping_reason", ""),
        "planner_notes": doc.get("planner_notes", ""),
        "locked": doc.get("locked", False),
        "planner_v3_crate_class": doc.get("planner_v3_crate_class"),
        "planner_v3_orientation": doc.get("planner_v3_orientation"),
        "planner_v3_splash_layers": doc.get("planner_v3_splash_layers"),
        "weight_band_status": doc.get("weight_band_status"),
        "planner_v3_pull_piece_ids": doc.get("planner_v3_pull_piece_ids") or [],
        "created_at": as_iso(doc.get("created_at")),
    }


def assignment_response(doc: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": doc["id"],
        "project_id": doc["project_id"],
        "piece_id": doc["piece_id"],
        "crate_id": doc["crate_id"],
        "assigned_at": as_iso(doc.get("assigned_at")),
    }


def pieces_grouped_by_crate(pieces: List[Dict[str, Any]], assignments: List[Dict[str, Any]]) -> Dict[int, List[Dict[str, Any]]]:
    pieces_by_id = {piece["id"]: piece for piece in pieces}
    grouped: Dict[int, List[Dict[str, Any]]] = {}
    for assignment in assignments:
        piece = pieces_by_id.get(assignment.get("piece_id"))
        crate_id = assignment.get("crate_id")
        if piece is None or crate_id is None:
            continue
        grouped.setdefault(crate_id, []).append(piece)
    for crate_pieces in grouped.values():
        crate_pieces.sort(key=lambda item: item["id"])
    return grouped


def project_planning_snapshot(project_id: int) -> Dict[str, Any]:
    project = projects_col.find_one({"id": project_id}, {"_id": 0})
    pieces = sorted(pieces_col.find({"project_id": project_id}, {"_id": 0}), key=lambda doc: doc["id"])
    crates = sorted(crates_col.find({"project_id": project_id}, {"_id": 0}), key=lambda doc: doc["id"])
    assignments = list(assignments_col.find({"project_id": project_id}, {"_id": 0}))
    return build_planning_snapshot(project, pieces, crates, assignments)


def clear_manual_container_plan(project_id: int) -> None:
    projects_col.update_one({"id": project_id}, {"$set": {"manual_container_plan": None}})


_THICKNESS_M = {"2CM": 0.02, "3CM": 0.03, "Mixed": 0.025}
_SQFT_TO_SQM = 0.0929
_WEIGHT_FACTORS_FALLBACK = {
    "Granite": {"2CM": 5.5, "3CM": 7.5, "Mixed": 6.5},
    "Quartz": {"2CM": 4.75, "3CM": 6.75, "Mixed": 5.75},
    "Marble": {"2CM": 6.0, "3CM": 8.0, "Mixed": 7.0},
    "Other": {"2CM": 5.5, "3CM": 7.5, "Mixed": 6.5},
}


def weight_factor(material: str, thickness: str, color: str = "") -> float:
    if color and material in COLOR_DENSITIES and color in COLOR_DENSITIES[material]:
        density = COLOR_DENSITIES[material][color]
        t_m = _THICKNESS_M.get(thickness, 0.025)
        return round(density * t_m * _SQFT_TO_SQM, 3)
    return _WEIGHT_FACTORS_FALLBACK.get(material, _WEIGHT_FACTORS_FALLBACK["Other"]).get(thickness, 6.5)


def calculate_edge_polish_machine(length: float, width: float, edge_area: str) -> float:
    try:
        l = float(length)
        w = float(width)
    except (TypeError, ValueError):
        return 0.0

    if l <= 0 or w <= 0 or not edge_area:
        return 0.0

    perimeter = 2 * (l + w)
    if edge_area in {"4 Sides", "Perimeter"}:
        return perimeter
    if edge_area == "3 Sides":
        return 2 * max(l, w) + min(l, w)
    if edge_area == "2 Sides":
        return 2 * max(l, w)
    if edge_area == "1 Side":
        return max(l, w)
    return 0.0


def piece_weight(piece: Dict[str, Any], material: str, thickness: str, color: str = "") -> float:
    sqft = (float(piece.get("length", 0)) * float(piece.get("width", 0))) / 144.0
    effective_thickness = piece.get("thickness") or thickness
    return sqft * weight_factor(material, effective_thickness, color) * int(piece.get("qty", 1))


def estimate_crate_dimensions(
    pieces: List[Dict[str, Any]],
    material: str,
    thickness: str,
    max_weight: float,
) -> Dict[str, Any]:
    if not pieces:
        return {
            "internal_length": 0.0,
            "internal_width": 0.0,
            "internal_height": 0.0,
            "external_length": 0.0,
            "external_width": 0.0,
            "external_height": 0.0,
            "sqft": 0.0,
            "weight": 0.0,
        }

    lengths = [float(piece.get("length", 0) or 0) for piece in pieces]
    widths = [float(piece.get("width", 0) or 0) for piece in pieces]
    total_sqft = sum(((float(piece.get("length", 0) or 0) * float(piece.get("width", 0) or 0)) / 144.0) * int(piece.get("qty", 1) or 1) for piece in pieces)
    total_weight = sum(piece_weight(piece, material, thickness) for piece in pieces)
    total_qty = sum(int(piece.get("qty", 1) or 1) for piece in pieces)

    longest_piece = max(lengths) if lengths else 0.0
    widest_piece = max(widths) if widths else 0.0

    # Internal sizing uses the largest piece plus handling clearance.
    clearance = 6.0
    internal_length = max(longest_piece + clearance, 0.0)
    internal_width = max(widest_piece + clearance, 0.0)

    # Height is a heuristic based on stack count and weight. Stone crates need
    # enough headroom for edge protectors, separators, and lifting access.
    height_from_qty = 18.0 + (total_qty * 1.25)
    height_from_weight = 18.0 + (total_weight / 75.0 if total_weight else 0.0)
    height_from_area = 18.0 + (total_sqft / 18.0 if total_sqft else 0.0)
    internal_height = max(24.0, height_from_qty, height_from_weight, height_from_area)
    internal_height = min(internal_height, 60.0)

    # External sizing adds crate walls, runners, and a little top/bottom slack.
    wall_allowance = 3.0
    height_allowance = 6.0
    external_length = internal_length + wall_allowance
    external_width = internal_width + wall_allowance
    external_height = internal_height + height_allowance

    return {
        "internal_length": round(internal_length, 1),
        "internal_width": round(internal_width, 1),
        "internal_height": round(internal_height, 1),
        "external_length": round(external_length, 1),
        "external_width": round(external_width, 1),
        "external_height": round(external_height, 1),
        "sqft": round(total_sqft, 2),
        "weight": round(total_weight, 2),
        "max_weight": max_weight,
    }


def crate_group_name(crate_name: str) -> str:
    if "-" not in crate_name:
        return crate_name
    return crate_name.rsplit("-", 1)[0]


def container_plan(
    total_weight: float,
    crate_count: int,
    average_utilization: float,
    distinct_families: int,
    distinct_destinations: int,
    underfilled_count: int,
) -> Dict[str, Any]:
    # Conservative payload guidance based on common dry-container specs from
    # Hapag-Lloyd and Maersk:
    # - 20ft dry: ~28.1t payload, ~33cbm
    # - 40ft dry: ~28.7t payload, ~67cbm
    # For stone, volume/handling usually pushes us to 40ft earlier than payload does.
    container_specs = {
        "20ft": {"payload_kg": 28130, "cbm": 33.2},
        "40ft": {"payload_kg": 28750, "cbm": 67.7},
    }

    if total_weight > container_specs["40ft"]["payload_kg"]:
        return {
            "recommended": "multiple",
            "booking_action": "Split into 2 containers",
            "reason": "This load is over a single standard dry container payload, so it should be split into multiple containers or multiple deliveries.",
            "alternatives": [
                f"20ft dry: up to {container_specs['20ft']['payload_kg']:,} kg and about {container_specs['20ft']['cbm']:.1f} cbm",
                f"40ft dry: up to {container_specs['40ft']['payload_kg']:,} kg and about {container_specs['40ft']['cbm']:.1f} cbm",
            ],
            "next_step": "Split by destination or by load sequence before choosing a carrier booking.",
        }

    if underfilled_count >= max(3, crate_count // 3) and total_weight <= 15000:
        return {
            "recommended": "consolidate",
            "booking_action": "Hold for consolidation",
            "reason": "Several crates are underfilled. Before booking a container, consolidate small crates so you avoid shipping air.",
            "alternatives": [
                "Merge low-fill crates that share the same family or destination.",
                f"20ft dry: up to {container_specs['20ft']['payload_kg']:,} kg and about {container_specs['20ft']['cbm']:.1f} cbm",
                f"40ft dry: up to {container_specs['40ft']['payload_kg']:,} kg and about {container_specs['40ft']['cbm']:.1f} cbm",
            ],
            "next_step": "Repack to raise crate fill to roughly 85-95% before picking the container.",
        }

    if total_weight <= 9000 and crate_count <= 8 and average_utilization >= 88 and distinct_families <= 3:
        recommended = "20ft"
        booking_action = "Book 1 x 20ft"
        reason = (
            "This looks like a compact, well-filled export load. A 20ft dry container is usually the better economical choice "
            "when the crate count is modest and the total weight stays comfortably below payload limits."
        )
    elif total_weight <= 12000 and crate_count <= 14 and distinct_families <= 4 and distinct_destinations <= 6:
        recommended = "40ft"
        booking_action = "Book 1 x 40ft"
        reason = (
            "This is a typical mixed countertop / sill export job. A 40ft dry container is usually the safer choice because "
            "it gives more floor length and less stacking pressure, even though the payload could fit in a 20ft."
        )
    elif total_weight <= 18000:
        recommended = "40ft"
        booking_action = "Book 1 x 40ft"
        reason = (
            "This is a mid-size countertop load. A 40ft dry container gives better floor space for mixed apartment drops, "
            "window sills, and crate staging, even though a 20ft could still handle the weight."
        )
    else:
        recommended = "40ft"
        booking_action = "Book 1 x 40ft"
        reason = (
            "This is a heavy single-container candidate. A 40ft dry container is the safer default for stone because it gives "
            "more loading flexibility and breathing room for padding, blocking, and access."
        )

    return {
        "recommended": recommended,
        "booking_action": booking_action,
        "reason": reason,
        "alternatives": [
            f"20ft dry: up to {container_specs['20ft']['payload_kg']:,} kg and about {container_specs['20ft']['cbm']:.1f} cbm",
            f"40ft dry: up to {container_specs['40ft']['payload_kg']:,} kg and about {container_specs['40ft']['cbm']:.1f} cbm",
        ],
        "next_step": "Keep the crate mix tight and avoid leaving 20-30% empty space in too many crates.",
    }


def crate_insights(project_id: int) -> Dict[str, Any]:
    snapshot = project_planning_snapshot(project_id)
    crate_rows = snapshot["crate_rows"]
    loading_plan = build_container_plan(
        crate_rows,
        manual_plan=snapshot["project"].get("manual_container_plan"),
        preferred_mode=snapshot["project"].get("preferred_container_mode"),
    )

    exception_rows = list(snapshot["exception_rows"])
    for container in loading_plan["containers"]:
        for warning in container.get("warnings", []):
            exception_rows.append(
                {
                    "scope": "container",
                    "id": container["id"],
                    "name": container["type"],
                    "severity": "yellow",
                    "message": warning,
                }
            )

    recommended_target = 92 if any(crate["efficiency_status"] == "red" for crate in crate_rows) else 97

    return {
        "summary": {
            "crates_created": len(crate_rows),
            "shipment_weight": snapshot["total_gross_weight"],
            "net_stone_weight": snapshot["total_weight"],
            "recommended_containers": loading_plan["recommendation"].get("booking_action", ""),
            "container_mode": loading_plan["recommendation"].get("mode_label", ""),
            "cost_index": loading_plan["recommendation"].get("cost_index", 0.0),
        },
        "total_weight": snapshot["total_weight"],
        "shipment_weight": snapshot["total_gross_weight"],
        "crate_count": len(crate_rows),
        "average_utilization": snapshot["average_fill_percent"],
        "average_weight_utilization": snapshot["average_gross_utilization"],
        "recommended_utilization_target": recommended_target,
        "adjusted_target_weight": round(
            snapshot["total_gross_weight"] / len(crate_rows), 0
        ) if crate_rows else 0,
        "distinct_families": snapshot["family_count"],
        "distinct_destinations": snapshot["destination_count"],
        "efficiency_kpis": {
            "average_fill_percent": snapshot["average_fill_percent"],
            "average_weight_utilization": snapshot["average_gross_utilization"],
            "average_payload_utilization": snapshot["average_payload_utilization"],
            "total_sqft": snapshot["total_sqft"],
            "piece_count": snapshot["piece_count"],
            "item_count": snapshot["item_count"],
        },
        "container_plan": loading_plan["recommendation"],
        "container_options": loading_plan["options"],
        "container_loading_plan": loading_plan,
        "crates": crate_rows,
        "underfilled_crates": snapshot["underfilled_crates"],
        "exceptions": exception_rows,
    }


class PieceCreate(BaseModel):
    part: str
    part_no: str = ""
    category: str
    drawing: str = ""
    length: float
    width: float
    thickness: str = "3CM"
    qty: int = 1
    unit: str = ""
    building: str = ""
    floor: str = ""
    flat: str = ""
    sink_type: str = "No Sink"
    sink_cut: str = "-"
    tap_holes: str = "-"
    grooves: str = "-"
    fragility: str = "Standard"
    orientation: str = "Auto"
    delivery_priority: str = "Standard"
    stack_preference: str = "Auto"
    weight_override: float = 0.0
    edge: str = "None"
    edge_area: str = ""
    edge_polish_machine: float = 0.0
    edge_map: Dict[str, str] = {}
    edge_polish_manual: str = ""
    radius: str = "-"
    radius_value: str = ""
    radius_corners: Dict[str, bool] = {}
    shape_type: str = ""
    notes: str = ""


class PieceUpdate(BaseModel):
    part: str = ""
    part_no: str = ""
    category: str = ""
    drawing: str = ""
    length: float = 0.0
    width: float = 0.0
    thickness: str = "3CM"
    qty: int = 1
    unit: str = ""
    building: str = ""
    floor: str = ""
    flat: str = ""
    sink_type: str = "No Sink"
    sink_cut: str = "-"
    tap_holes: str = "-"
    grooves: str = "-"
    fragility: str = "Standard"
    orientation: str = "Auto"
    delivery_priority: str = "Standard"
    stack_preference: str = "Auto"
    weight_override: float = 0.0
    edge: str = "None"
    edge_area: str = ""
    edge_polish_machine: float = 0.0
    edge_map: Dict[str, str] = {}
    edge_polish_manual: str = ""
    radius: str = "-"
    radius_value: str = ""
    radius_corners: Dict[str, bool] = {}
    shape_type: str = ""
    notes: str = ""


class ProjectUpdate(BaseModel):
    name: str
    material: str
    thickness: str
    stone_color: str = ""
    crate_wood_type: str = "Pine"
    crate_wood_thickness: float = 1.25
    preferred_container_mode: str = "recommended"
    customer: str
    job_number: str
    date: str
    flat_format: str = "3-digit"
    description_thickness_map: Dict[str, str] = {}


class PlannerPayloadUpdate(BaseModel):
    """Planner-only: 20ft payload cap (kg). Default 24,000; port unload commonly 28,000."""
    delivery_payload_cap_kg: float = 24000.0


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "https://stonedesk-wwrc.vercel.app",
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Perf-Validation", "X-Perf-Mongo", "X-Perf-Backend-Total"],
)


@app.get("/")
def root():
    return {"status": "online", "message": "StoneDesk API is running"}


@app.get("/api/projects/")
def get_projects():
    projects = sorted(projects_col.find({}, {"_id": 0}), key=lambda doc: doc["id"])
    return [project_response(p) for p in projects]


@app.post("/api/projects/")
def create_project():
    project_id = next_sequence("project")
    now = utc_now()
    doc = {
        "id": project_id,
        "name": "",
        "material": "Granite",
        "thickness": "3CM",
        "stone_color": "",
        "crate_wood_type": "Pine",
        "crate_wood_thickness": 1.25,
        "preferred_container_mode": "recommended",
        "customer": "",
        "job_number": "",
        "date": date.today().isoformat(),
        "created_at": now,
        "updated_at": now,
    }
    projects_col.insert_one(doc)
    return project_response(doc)


@app.get("/api/projects/{project_id}")
def get_project(project_id: int):
    project = projects_col.find_one({"id": project_id}, {"_id": 0})
    return project_response(project)


@app.patch("/api/projects/{project_id}/planner-payload")
def patch_planner_payload(project_id: int, data: PlannerPayloadUpdate):
    """Does not clear container plan — safe for planner toolbar."""
    cap = max(20000.0, min(32000.0, float(data.delivery_payload_cap_kg)))
    projects_col.update_one(
        {"id": project_id},
        {"$set": {"delivery_payload_cap_kg": cap, "updated_at": utc_now()}},
    )
    return {"message": "ok", "delivery_payload_cap_kg": cap}


@app.put("/api/projects/{project_id}")
def update_project(project_id: int, data: ProjectUpdate):
    update = {
        "name": data.name,
        "material": data.material,
        "thickness": data.thickness,
        "stone_color": data.stone_color,
        "crate_wood_type": data.crate_wood_type,
        "crate_wood_thickness": data.crate_wood_thickness,
        "preferred_container_mode": data.preferred_container_mode,
        "customer": data.customer,
        "job_number": data.job_number,
        "date": data.date,
        "flat_format": data.flat_format,
        "description_thickness_map": data.description_thickness_map,
        "updated_at": utc_now(),
    }
    projects_col.update_one({"id": project_id}, {"$set": update})
    clear_manual_container_plan(project_id)
    return {"message": "ok"}


VALID_PROJECT_STATUSES = {
    "draft", "review_pending", "approved_for_packing",
    "crate_planned", "packing_approved", "container_planned"
}


@app.patch("/api/projects/{project_id}/status")
def update_project_status(project_id: int, data: Dict[str, Any]):
    status = data.get("status", "")
    if status not in VALID_PROJECT_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status '{status}'")
    projects_col.update_one({"id": project_id}, {"$set": {"status": status, "updated_at": utc_now()}})
    return {"message": "ok", "status": status}


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int):
    pieces = list(pieces_col.find({"project_id": project_id}, {"id": 1}))
    piece_ids = [piece["id"] for piece in pieces]
    crates = list(crates_col.find({"project_id": project_id}, {"id": 1}))
    crate_ids = [crate["id"] for crate in crates]

    if piece_ids:
        assignments_col.delete_many({"project_id": project_id, "piece_id": {"$in": piece_ids}})
    if crate_ids:
        assignments_col.delete_many({"project_id": project_id, "crate_id": {"$in": crate_ids}})

    pieces_col.delete_many({"project_id": project_id})
    crates_col.delete_many({"project_id": project_id})
    projects_col.delete_one({"id": project_id})
    return {"message": "ok"}


@app.get("/api/projects/{project_id}/pieces/")
def get_pieces(project_id: int):
    pieces = sorted(pieces_col.find({"project_id": project_id}, {"_id": 0}), key=lambda doc: doc["id"])
    return [piece_response(piece) for piece in pieces]


@app.post("/api/projects/{project_id}/pieces/")
def create_piece(project_id: int, piece: PieceCreate):
    piece_id = next_sequence("piece")
    edge_polish_machine = piece.edge_polish_machine or calculate_edge_polish_machine(piece.length, piece.width, piece.edge_area)
    doc = {
        "id": piece_id,
        "project_id": project_id,
        "part": piece.part,
        "category": piece.category,
        "drawing": piece.drawing,
        "length": piece.length,
        "width": piece.width,
        "thickness": piece.thickness,
        "qty": piece.qty,
        "unit": piece.unit,
        "building": piece.building,
        "floor": piece.floor,
        "flat": piece.flat,
        "sink_type": piece.sink_type,
        "sink_cut": piece.sink_cut,
        "tap_holes": piece.tap_holes,
        "grooves": piece.grooves,
        "fragility": piece.fragility,
        "orientation": piece.orientation,
        "delivery_priority": piece.delivery_priority,
        "stack_preference": piece.stack_preference,
        "weight_override": piece.weight_override,
        "edge": piece.edge,
        "edge_area": piece.edge_area,
        "edge_polish_machine": edge_polish_machine,
        "radius": piece.radius,
        "notes": piece.notes,
        "created_at": utc_now(),
    }
    pieces_col.insert_one(doc)
    clear_manual_container_plan(project_id)
    return piece_response(doc)


@app.post("/api/projects/{project_id}/pieces/batch")
def create_pieces_batch(project_id: int, pieces_data: List[PieceCreate]):
    docs = []
    for piece in pieces_data:
        piece_id = next_sequence("piece")
        edge_polish_machine = piece.edge_polish_machine or calculate_edge_polish_machine(piece.length, piece.width, piece.edge_area)
        docs.append(
            {
                "id": piece_id,
                "project_id": project_id,
                "part": piece.part,
                "part_no": piece.part_no,
                "category": piece.category,
                "drawing": piece.drawing,
                "length": piece.length,
                "width": piece.width,
                "thickness": piece.thickness,
                "qty": piece.qty,
                "unit": piece.unit,
                "building": piece.building,
                "floor": piece.floor,
                "flat": piece.flat,
                "sink_type": piece.sink_type,
                "sink_cut": piece.sink_cut,
                "tap_holes": piece.tap_holes,
                "grooves": piece.grooves,
                "fragility": piece.fragility,
                "orientation": piece.orientation,
                "delivery_priority": piece.delivery_priority,
                "stack_preference": piece.stack_preference,
                "weight_override": piece.weight_override,
                "edge": piece.edge,
                "edge_area": piece.edge_area,
                "edge_polish_machine": edge_polish_machine,
                "edge_map": piece.edge_map,
                "edge_polish_manual": piece.edge_polish_manual,
                "radius": piece.radius,
                "radius_value": piece.radius_value,
                "radius_corners": piece.radius_corners,
                "shape_type": piece.shape_type,
                "notes": piece.notes,
                "created_at": utc_now(),
            }
        )
    if docs:
        pieces_col.insert_many(docs)
        clear_manual_container_plan(project_id)
    return {"message": f"Created {len(docs)} pieces"}


@app.get("/api/projects/{project_id}/drawings/")
def get_project_drawings(project_id: int):
    pieces = list(pieces_col.find({"project_id": project_id}, {"_id": 0}))
    drawings_map: dict = {}
    for piece in pieces:
        dn = piece.get("drawing") or "Unnamed"
        if dn not in drawings_map:
            drawings_map[dn] = {
                "drawing": dn,
                "category": piece.get("category") or "",
                "unit": piece.get("unit") or "",
                "fragility": piece.get("fragility") or "Standard",
                "orientation": piece.get("orientation") or "Auto",
                "delivery_priority": piece.get("delivery_priority") or "Standard",
                "stack_preference": piece.get("stack_preference") or "Auto",
                "weight_override": piece.get("weight_override") or 0,
                "pieces": [],
            }
        drawings_map[dn]["pieces"].append(piece)

    result = []
    for dn, data in drawings_map.items():
        pcs = data["pieces"]

        # Unique part rows (grouped by part_no + part + dims)
        seen: dict = {}
        for p in pcs:
            k = (p.get("part_no") or "", p.get("part") or "",
                 p.get("length") or 0, p.get("width") or 0)
            if k not in seen:
                seen[k] = {
                    "part_no": p.get("part_no") or "",
                    "part": p.get("part") or "",
                    "length": p.get("length") or 0,
                    "width": p.get("width") or 0,
                    "thickness": p.get("thickness") or "3CM",
                    "qty": 1,
                    "sink_type": p.get("sink_type") or "No Sink",
                    "sink_cut": p.get("sink_cut") or "-",
                    "tap_holes": p.get("tap_holes") or "-",
                    "grooves": p.get("grooves") or "-",
                    "edge": p.get("edge") or "None",
                    "edge_area": p.get("edge_area") or "",
                    "edge_map": p.get("edge_map") or {},
                    "edge_polish_manual": p.get("edge_polish_manual") or "",
                    "radius": p.get("radius") or "-",
                    "radius_value": p.get("radius_value") or "",
                    "radius_corners": p.get("radius_corners") or {},
                    "shape_type": p.get("shape_type") or "",
                    "notes": p.get("notes") or "",
                }
            else:
                seen[k]["qty"] += 1

        # Destination matrix
        buildings = sorted({str(p.get("building") or "") for p in pcs if p.get("building")})
        floors = sorted({str(p.get("floor") or "") for p in pcs if p.get("floor")})
        cells: dict = {}
        for p in pcs:
            b = str(p.get("building") or "")
            fl = str(p.get("floor") or "")
            ft = str(p.get("flat") or "")
            if b and fl and ft:
                key = f"{b}__{fl}"
                if key not in cells:
                    cells[key] = []
                if not any(e["flat"] == ft for e in cells[key]):
                    cells[key].append({"flat": ft, "qty": 1})

        result.append({
            "drawing": dn,
            "category": data["category"],
            "unit": data["unit"],
            "fragility": data["fragility"],
            "orientation": data["orientation"],
            "delivery_priority": data["delivery_priority"],
            "stack_preference": data["stack_preference"],
            "weight_override": data["weight_override"],
            "piece_count": len(pcs),
            "unique_parts": list(seen.values()),
            "buildings": buildings,
            "floors": floors,
            "cells": cells,
            "destination_summary": sorted({
                "/".join(filter(None, [str(p.get("building") or ""), str(p.get("floor") or ""), str(p.get("flat") or "")]))
                for p in pcs
            }),
        })

    result.sort(key=lambda d: d["drawing"])
    return result


@app.delete("/api/projects/{project_id}/pieces/")
def delete_all_pieces(project_id: int):
    pieces = list(pieces_col.find({"project_id": project_id}, {"id": 1}))
    piece_ids = [piece["id"] for piece in pieces]
    if piece_ids:
        assignments_col.delete_many({"project_id": project_id, "piece_id": {"$in": piece_ids}})
    pieces_col.delete_many({"project_id": project_id})
    clear_manual_container_plan(project_id)
    return {"message": "ok"}


@app.delete("/api/pieces/{piece_id}")
def delete_piece(piece_id: int):
    piece = pieces_col.find_one({"id": piece_id}, {"project_id": 1})
    if piece:
        assignments_col.delete_many({"project_id": piece["project_id"], "piece_id": piece_id})
        clear_manual_container_plan(piece["project_id"])
    pieces_col.delete_one({"id": piece_id})
    return {"message": "ok"}


@app.put("/api/pieces/{piece_id}")
def update_piece(piece_id: int, piece: PieceUpdate):
    existing = pieces_col.find_one({"id": piece_id}, {"project_id": 1})
    if not existing:
        return {"message": "ok"}

    edge_polish_machine = piece.edge_polish_machine or calculate_edge_polish_machine(piece.length, piece.width, piece.edge_area)
    update = {
        "part": piece.part,
        "part_no": piece.part_no,
        "category": piece.category,
        "drawing": piece.drawing,
        "length": piece.length,
        "width": piece.width,
        "thickness": piece.thickness,
        "qty": piece.qty,
        "unit": piece.unit,
        "building": piece.building,
        "floor": piece.floor,
        "flat": piece.flat,
        "sink_type": piece.sink_type,
        "sink_cut": piece.sink_cut,
        "tap_holes": piece.tap_holes,
        "grooves": piece.grooves,
        "fragility": piece.fragility,
        "orientation": piece.orientation,
        "delivery_priority": piece.delivery_priority,
        "stack_preference": piece.stack_preference,
        "weight_override": piece.weight_override,
        "edge": piece.edge,
        "edge_area": piece.edge_area,
        "edge_polish_machine": edge_polish_machine,
        "edge_map": piece.edge_map,
        "edge_polish_manual": piece.edge_polish_manual,
        "radius": piece.radius,
        "radius_value": piece.radius_value,
        "radius_corners": piece.radius_corners,
        "shape_type": piece.shape_type,
        "notes": piece.notes,
    }
    pieces_col.update_one({"id": piece_id}, {"$set": update})
    clear_manual_container_plan(existing["project_id"])
    return {"message": "ok"}


def _fmt_edge_map(edge_map) -> str:
    """Compact per-side edge summary: 'T:Polish, B:Cut, L:None, R:Polish'"""
    if not edge_map:
        return ""
    sides = [("top", "T"), ("bottom", "B"), ("left", "L"), ("right", "R")]
    return ", ".join(f"{lbl}:{(edge_map.get(s) or 'none').capitalize()}" for s, lbl in sides)


def _fmt_radius_corners(radius_corners) -> str:
    """Active corner abbreviations: 'TL, BR'"""
    if not radius_corners:
        return ""
    labels = {"top_left": "TL", "top_right": "TR", "bottom_left": "BL", "bottom_right": "BR"}
    active = [labels[k] for k in ("top_left", "top_right", "bottom_left", "bottom_right") if radius_corners.get(k)]
    return ", ".join(active) if active else "—"


def _bold_row(ws, row_num: int) -> None:
    try:
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill("solid", fgColor="1E293B")
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
    except Exception:
        pass


def _hide_empty_columns(ws, header_row: int = 1) -> None:
    """Hide columns that have no non-empty/non-zero data rows after the header."""
    try:
        from openpyxl.utils import get_column_letter
        max_row = ws.max_row
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            has_value = False
            for row_idx in range(header_row + 1, max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and val != '' and val != 0 and val != 0.0:
                    has_value = True
                    break
            if not has_value:
                ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    except Exception:
        pass


def crate_serial_for_project(project_id: int) -> int:
    last_crate = crates_col.find_one(
        {"project_id": project_id},
        sort=[("id", -1)],
        projection={"crate_id": 1},
    )
    if not last_crate:
        return 1

    crate_id = str(last_crate.get("crate_id", ""))
    if crate_id.startswith("CR"):
        try:
            return int(crate_id[2:]) + 1
        except ValueError:
            return 1
    return 1


@app.get("/api/projects/{project_id}/crates/")
def get_crates(project_id: int):
    snapshot = project_planning_snapshot(project_id)
    return snapshot["crate_rows"]


@app.post("/api/projects/{project_id}/crates/")
def create_crate(project_id: int, data: Dict[str, Any]):
    serial = crate_serial_for_project(project_id)
    piece_ids = [int(piece_id) for piece_id in data.get("piece_ids", []) if str(piece_id).strip()]
    internal_length = float(data.get("internal_length", 0) or 0)
    internal_width = float(data.get("internal_width", 0) or 0)
    internal_height = float(data.get("internal_height", 0) or 0)
    external_length = float(data.get("external_length", 0) or 0)
    external_width = float(data.get("external_width", 0) or 0)
    external_height = float(data.get("external_height", 0) or 0)
    dimension_mode = data.get("dimension_mode") or (
        "manual" if any([internal_length, internal_width, internal_height, external_length, external_width, external_height]) else "auto"
    )
    crate = {
        "id": next_sequence("crate"),
        "project_id": project_id,
        "crate_id": f"CR{serial:04d}",
        "name": data.get("name") or f"Crate {serial}",
        "max_weight": float(data.get("max_weight", 1000) or 1000),
        "locked": bool(data.get("locked", False)),
        "custom": bool(data.get("custom", bool(piece_ids))),
        "reserved_space_pct": float(data.get("reserved_space_pct", 0) or 0),
        "planner_notes": str(data.get("planner_notes", "") or ""),
        "dimension_mode": dimension_mode,
        "stackable": data.get("stackable") if "stackable" in data else None,
        "forklift_entry": data.get("forklift_entry"),
        "reinforcement": data.get("reinforcement") if "reinforcement" in data else None,
        "wood_thickness": float(data.get("wood_thickness", 0) or 0),
        "internal_length": internal_length,
        "internal_width": internal_width,
        "internal_height": internal_height,
        "external_length": external_length,
        "external_width": external_width,
        "external_height": external_height,
        "sqft": float(data.get("sqft", 0) or 0),
        "weight": float(data.get("weight", 0) or 0),
        "packing_family": str(data.get("packing_family", "") or ""),
        "created_at": utc_now(),
    }
    crates_col.insert_one(crate)

    if piece_ids:
        assignments_col.delete_many({"project_id": project_id, "piece_id": {"$in": piece_ids}})
        assignments_col.insert_many(
            [
                {
                    "id": next_sequence("assignment"),
                    "project_id": project_id,
                    "piece_id": piece_id,
                    "crate_id": crate["id"],
                    "assigned_at": utc_now(),
                }
                for piece_id in piece_ids
            ]
        )
    clear_manual_container_plan(project_id)
    return {"id": crate["id"], "crate_id": crate["crate_id"]}


@app.put("/api/projects/{project_id}/crates/{crate_id}")
def update_crate(project_id: int, crate_id: int, data: Dict[str, Any]):
    existing = crates_col.find_one({"project_id": project_id, "id": crate_id}, {"_id": 0})
    if not existing:
        return {"message": "ok"}

    was_locked = bool(existing.get("locked"))
    unlocking = was_locked and data.get("locked") is False

    if was_locked and not unlocking:
        if data.get("reset_dimensions"):
            raise HTTPException(status_code=400, detail="Unlock the crate before resetting dimensions.")
        dim_fields = [
            "internal_length",
            "internal_width",
            "internal_height",
            "external_length",
            "external_width",
            "external_height",
        ]
        if any(f in data for f in dim_fields):
            raise HTTPException(status_code=400, detail="Unlock the crate before editing dimensions.")
        if "dimension_mode" in data and (data.get("dimension_mode") or "") != (existing.get("dimension_mode") or ""):
            raise HTTPException(status_code=400, detail="Unlock the crate before changing dimension mode.")

    update: Dict[str, Any] = {
        "name": data.get("name", existing.get("name", existing.get("crate_id", "Crate"))),
        "max_weight": float(data.get("max_weight", existing.get("max_weight", 1000)) or existing.get("max_weight", 1000)),
        "locked": bool(data.get("locked", existing.get("locked", False))),
        "custom": bool(data.get("custom", existing.get("custom", False))),
        "reserved_space_pct": float(data.get("reserved_space_pct", existing.get("reserved_space_pct", 0)) or 0),
        "planner_notes": str(data.get("planner_notes", existing.get("planner_notes", "")) or ""),
        "updated_at": utc_now(),
    }

    if "stackable" in data:
        update["stackable"] = data.get("stackable")
    if "forklift_entry" in data:
        update["forklift_entry"] = data.get("forklift_entry")
    if "reinforcement" in data:
        update["reinforcement"] = data.get("reinforcement")
    if "wood_thickness" in data:
        update["wood_thickness"] = float(data.get("wood_thickness", existing.get("wood_thickness", 0)) or 0)

    if data.get("reset_dimensions"):
        update.update(
            {
                "dimension_mode": "auto",
                "internal_length": 0.0,
                "internal_width": 0.0,
                "internal_height": 0.0,
                "external_length": 0.0,
                "external_width": 0.0,
                "external_height": 0.0,
            }
        )
    else:
        dimension_fields = [
            "internal_length",
            "internal_width",
            "internal_height",
            "external_length",
            "external_width",
            "external_height",
        ]
        manual_dimension_provided = False
        for field in dimension_fields:
            if field in data:
                update[field] = float(data.get(field, existing.get(field, 0)) or 0)
                manual_dimension_provided = True
        if "dimension_mode" in data:
            update["dimension_mode"] = data.get("dimension_mode") or existing.get("dimension_mode", "auto")
        elif manual_dimension_provided:
            update["dimension_mode"] = "manual"

    crates_col.update_one({"project_id": project_id, "id": crate_id}, {"$set": update})
    clear_manual_container_plan(project_id)
    return {"message": "ok"}


@app.post("/api/projects/{project_id}/crates/merge")
def merge_crates(project_id: int, data: Dict[str, Any]):
    crate_ids = [int(crate_id) for crate_id in data.get("crate_ids", []) if str(crate_id).strip()]
    if len(crate_ids) < 2:
        return {"message": "select at least two crates"}

    target_id = int(data.get("target_crate_id") or crate_ids[0])
    target = crates_col.find_one({"project_id": project_id, "id": target_id}, {"_id": 0})
    if not target:
        return {"message": "target not found"}

    involved = list(dict.fromkeys(crate_ids + [target_id]))
    for cid in involved:
        row = crates_col.find_one({"project_id": project_id, "id": cid}, {"locked": 1})
        if row and row.get("locked"):
            raise HTTPException(status_code=400, detail="Merge blocked — a selected crate is locked.")

    source_ids = [crate_id for crate_id in crate_ids if crate_id != target_id]
    assignments = list(assignments_col.find({"project_id": project_id, "crate_id": {"$in": source_ids}}, {"_id": 0}))
    for assignment in assignments:
        assignments_col.update_one(
            {"id": assignment["id"]},
            {"$set": {"crate_id": target_id, "assigned_at": utc_now()}},
        )

    if data.get("name"):
        crates_col.update_one(
            {"project_id": project_id, "id": target_id},
            {"$set": {"name": data["name"], "updated_at": utc_now()}},
        )

    crates_col.delete_many({"project_id": project_id, "id": {"$in": source_ids}})
    clear_manual_container_plan(project_id)
    return {"message": "ok"}


@app.post("/api/projects/{project_id}/crates/split")
def split_crate(project_id: int, data: Dict[str, Any]):
    piece_ids = [int(piece_id) for piece_id in data.get("piece_ids", []) if str(piece_id).strip()]
    if not piece_ids:
        return {"message": "no pieces selected"}

    a0 = assignments_col.find_one({"project_id": project_id, "piece_id": piece_ids[0]})
    if a0:
        src_c = crates_col.find_one({"project_id": project_id, "id": a0["crate_id"]}, {"locked": 1})
        if src_c and src_c.get("locked"):
            raise HTTPException(status_code=400, detail="Cannot split a locked crate.")

    serial = crate_serial_for_project(project_id)
    crate_doc = {
        "id": next_sequence("crate"),
        "project_id": project_id,
        "crate_id": f"CR{serial:04d}",
        "name": data.get("name") or f"Crate {serial}",
        "max_weight": float(data.get("max_weight", 1000) or 1000),
        "locked": bool(data.get("locked", False)),
        "custom": True,
        "reserved_space_pct": float(data.get("reserved_space_pct", 0) or 0),
        "planner_notes": str(data.get("planner_notes", "") or ""),
        "dimension_mode": data.get("dimension_mode", "auto") or "auto",
        "internal_length": float(data.get("internal_length", 0) or 0),
        "internal_width": float(data.get("internal_width", 0) or 0),
        "internal_height": float(data.get("internal_height", 0) or 0),
        "external_length": float(data.get("external_length", 0) or 0),
        "external_width": float(data.get("external_width", 0) or 0),
        "external_height": float(data.get("external_height", 0) or 0),
        "sqft": 0.0,
        "weight": 0.0,
        "created_at": utc_now(),
    }
    crates_col.insert_one(crate_doc)

    assignments_col.delete_many({"project_id": project_id, "piece_id": {"$in": piece_ids}})
    assignments_col.insert_many(
        [
            {
                "id": next_sequence("assignment"),
                "project_id": project_id,
                "piece_id": piece_id,
                "crate_id": crate_doc["id"],
                "assigned_at": utc_now(),
            }
            for piece_id in piece_ids
        ]
    )
    clear_manual_container_plan(project_id)
    return {"id": crate_doc["id"], "crate_id": crate_doc["crate_id"]}


@app.get("/api/projects/{project_id}/dispatch-hierarchy")
def get_dispatch_hierarchy(project_id: int):
    """Returns building → floor → flat hierarchy discovered from project pieces."""
    pieces = list(pieces_col.find({"project_id": project_id}, {"_id": 0}))
    return discover_dispatch_hierarchy(pieces)


@app.get("/api/projects/{project_id}/families")
def get_packing_families(project_id: int):
    """
    Returns computed packing families for the project, annotated with current
    crate assignments so the family builder UI can show assigned/unassigned state.
    """
    from collections import Counter

    pieces = list(pieces_col.find({"project_id": project_id}, {"_id": 0}))
    if not pieces:
        return []

    project_doc = projects_col.find_one({"id": project_id}, {"_id": 0})
    material = (project_doc or {}).get("material", "Granite")
    thickness = (project_doc or {}).get("thickness", "3CM")

    # piece_id → crate_id
    raw_assignments = assignments_col.find({"project_id": project_id}, {"_id": 0, "piece_id": 1, "crate_id": 1})
    piece_to_crate: Dict[int, int] = {a["piece_id"]: a["crate_id"] for a in raw_assignments}

    # crate_id → crate_id string (for display)
    crate_docs = {c["id"]: c.get("crate_id", "") for c in crates_col.find({"project_id": project_id}, {"_id": 0, "id": 1, "crate_id": 1})}

    families = build_packing_families(pieces)
    result = []
    for fam in families:
        all_pieces = fam["all_pieces"]
        piece_ids = [p["id"] for p in all_pieces]

        # Determine primary crate assignment (most common)
        assigned_crate_ids = [piece_to_crate[pid] for pid in piece_ids if pid in piece_to_crate]
        if assigned_crate_ids:
            crate_counts = Counter(assigned_crate_ids)
            primary_db_id = crate_counts.most_common(1)[0][0]
            is_split = len(crate_counts) > 1
        else:
            primary_db_id = None
            is_split = False

        total_weight = sum(det_piece_weight(p, material, thickness) for p in all_pieces)
        cat_cfg = DET_CATEGORY_CONFIG.get(fam["category"], DET_CATEGORY_CONFIG["misc"])

        # Extract location from first piece (all pieces in a family share flat/floor/building)
        ref = all_pieces[0] if all_pieces else {}
        building = str(ref.get("building", "") or "").strip()
        floor = str(ref.get("floor", "") or "").strip()
        flat = str(ref.get("flat", "") or "").strip()

        result.append({
            "family_id": fam["family_id"],
            "flat_key": fam["flat_key"],
            "building": building,
            "floor": floor,
            "flat": flat,
            "category": fam["category"],
            "category_label": cat_cfg["label"],
            "main_piece_ids": [p["id"] for p in fam["main_pieces"]],
            "splash_piece_ids": [p["id"] for p in fam["splash_pieces"]],
            "all_piece_ids": piece_ids,
            "main_part_nos": [str(p.get("part_no", "") or "") for p in fam["main_pieces"]],
            "splash_part_nos": [str(p.get("part_no", "") or "") for p in fam["splash_pieces"]],
            "main_count": len(fam["main_pieces"]),
            "splash_count": len(fam["splash_pieces"]),
            "total_pieces": len(all_pieces),
            "total_weight_kg": round(total_weight, 1),
            "current_crate_db_id": primary_db_id,
            "current_crate_label": crate_docs.get(primary_db_id, "") if primary_db_id else None,
            "is_split": is_split,
            "ideal_lo": cat_cfg["ideal_lo"],
            "ideal_hi": cat_cfg["ideal_hi"],
            "max_kg": cat_cfg["max_kg"],
        })

    return result


@app.post("/api/projects/{project_id}/crates/auto-generate")
def auto_generate(project_id: int, data: Dict[str, Any]):
    pieces = list(pieces_col.find({"project_id": project_id}, {"_id": 0}))
    if not pieces:
        return {"message": "no pieces"}

    project = projects_col.find_one({"id": project_id}, {"_id": 0})
    if not project:
        return {"message": "project not found"}

    status = project.get("status") or "draft"
    if status not in {"approved_for_packing", "crate_planned", "container_planned"}:
        raise HTTPException(
            status_code=400,
            detail="Approve the project for packing before generating a crate plan.",
        )

    existing_crates = list(crates_col.find({"project_id": project_id}, {"_id": 0}))
    locked_crates = [crate for crate in existing_crates if crate.get("locked")]
    locked_crate_ids = [crate["id"] for crate in locked_crates]
    locked_assignments = list(
        assignments_col.find({"project_id": project_id, "crate_id": {"$in": locked_crate_ids}}, {"_id": 0})
    ) if locked_crate_ids else []
    locked_piece_ids = {assignment["piece_id"] for assignment in locked_assignments}

    unlocked_crate_ids = [crate["id"] for crate in existing_crates if crate["id"] not in locked_crate_ids]
    if unlocked_crate_ids:
        assignments_col.delete_many({"project_id": project_id, "crate_id": {"$in": unlocked_crate_ids}})
        crates_col.delete_many({"project_id": project_id, "id": {"$in": unlocked_crate_ids}})

    _stone_color = project.get("stone_color", "") or ""
    _mat = project.get("material", "Granite")
    _thick = project.get("thickness", "3CM")
    unlocked_pieces = [p for p in pieces if p["id"] not in locked_piece_ids]

    dispatch_selection = data.get("dispatch_selection") or project.get("dispatch_selection") or {}
    if dispatch_selection:
        projects_col.update_one(
            {"id": project_id},
            {"$set": {"dispatch_selection": dispatch_selection, "updated_at": utc_now()}},
        )

    v3 = run_v3_planner(unlocked_pieces, project, dispatch_selection)
    specs = v3["crates"]
    layout = v3.get("container_layout") or {}
    container_plan_list = v3.get("containers") or ([layout] if layout.get("placements") is not None else [])

    crate_docs: List[Dict[str, Any]] = []
    assignment_docs: List[Dict[str, Any]] = []
    next_serial = crate_serial_for_project(project_id)
    _wood_thick = float(project.get("crate_wood_thickness", 1.5) or 1.5)

    for dispatch_seq, spec in enumerate(specs, start=1):
        dims = spec["dimensions"]
        sqft = sum(
            (float(p.get("length", 0) or 0) * float(p.get("width", 0) or 0) / 144.0)
            * max(1, int(p.get("qty", 1) or 1))
            for p in spec["pieces"]
        )
        wt = float(spec["total_weight_kg"])

        crate_doc = {
            "id": next_sequence("crate"),
            "project_id": project_id,
            "crate_id": f"CR{next_serial:04d}",
            "name": spec["name"],
            "max_weight": spec["max_weight"],
            "locked": False,
            "custom": False,
            "reserved_space_pct": 0.0,
            "planner_notes": "; ".join(spec.get("warnings") or []),
            "dimension_mode": "auto",
            "stackable": spec.get("orientation") == "horizontal",
            "forklift_entry": None,
            "reinforcement": None,
            "wood_type": project.get("crate_wood_type", "Pine"),
            "wood_thickness": dims.get("wood_thickness", _wood_thick),
            "internal_length": dims["internal_length"],
            "internal_width": dims["internal_width"],
            "internal_height": dims["internal_height"],
            "external_length": dims["external_length"],
            "external_width": dims["external_width"],
            "external_height": dims["external_height"],
            "sqft": round(sqft, 2),
            "weight": round(wt, 2),
            "crate_type": spec["crate_type_label"],
            "packing_mode": "v3",
            "primary_flat": spec["dispatch_group"],
            "secondary_flats": [],
            "weight_band_status": spec.get("weight_band_status", ""),
            "grouping_reason": spec.get("grouping_reason", ""),
            "packing_family": spec.get("category", ""),
            "splash_layer": bool(spec.get("splash_layer")),
            "main_layer_piece_ids": spec.get("main_layer_piece_ids", []),
            "splash_layer_piece_ids": spec.get("splash_layer_piece_ids", []),
            "packing_warnings": spec.get("warnings", []),
            "planner_v3_crate_class": spec.get("crate_class"),
            "planner_v3_orientation": spec.get("orientation"),
            "planner_v3_splash_layers": spec.get("splash_layers", []),
            "planner_v3_pull_piece_ids": spec.get("pull_candidate_piece_ids") or [],
            "dispatch_order": dispatch_seq,
            "created_at": utc_now(),
        }
        next_serial += 1
        crate_docs.append(crate_doc)
        for piece in spec["pieces"]:
            assignment_docs.append({
                "id": next_sequence("assignment"),
                "project_id": project_id,
                "piece_id": piece["id"],
                "crate_id": crate_doc["id"],
                "assigned_at": utc_now(),
            })

    if crate_docs:
        crates_col.insert_many(crate_docs)
    if assignment_docs:
        assignments_col.insert_many(assignment_docs)

    idx_to_crate_id = {i: doc["crate_id"] for i, doc in enumerate(crate_docs)}

    manual_plan_containers: List[Dict[str, Any]] = []
    enriched_layouts: List[Dict[str, Any]] = []
    for ci, cont in enumerate(container_plan_list):
        if not cont:
            continue
        pls = cont.get("placements") or []
        manual_placements: List[Dict[str, Any]] = []
        sorted_pls = linear_manual_sort_placements(pls, crate_docs)
        for order_idx, pl in enumerate(sorted_pls, start=1):
            cid = idx_to_crate_id.get(pl.get("crate_index"))
            if not cid:
                continue
            manual_placements.append({
                "crate_id": cid,
                "x": float(pl.get("x", 0) or 0),
                "y": float(pl.get("y", 0) or 0),
                "rotated": bool(pl.get("rotated", False)),
                "stack_level": int(pl.get("stack_level", 0) or 0),
                "loading_order": order_idx,
                "unload_order": max(1, len(sorted_pls) - order_idx + 1),
            })
        ctype = str(cont.get("type") or cont.get("container_type") or "20ft")
        manual_plan_containers.append({
            "id": f"V3-{ctype.upper()}-{project_id}-{ci + 1}",
            "type": ctype,
            "container_id": cont.get("container_id"),
            "placements": manual_placements,
        })
        if crate_docs:
            enriched_layouts.append(enrich_layout_with_crates(cont, crate_docs))
        else:
            enriched_layouts.append(cont)

    manual_plan = {"containers": manual_plan_containers}

    layout_persist = enriched_layouts[0] if enriched_layouts else (enrich_layout_with_crates(layout, crate_docs) if crate_docs else layout)

    projects_col.update_one(
        {"id": project_id},
        {
            "$set": {
                "manual_container_plan": manual_plan,
                "planner_v3_layout": layout_persist,
                "planner_v3_containers": enriched_layouts,
                "planner_v3_summary": v3.get("summary") or {},
                "planner_v3_container_optimization": v3.get("container_optimization") or {},
                "status": "crate_planned",
                "updated_at": utc_now(),
            }
        },
    )

    return {
        "message": f"Created {len(crate_docs)} crates",
        "locked_preserved": len(locked_crates),
        "planner_v3": {
            "container": layout_persist,
            "containers": enriched_layouts,
            "summary": v3.get("summary") or {},
            "suggest_40ft": any(c.get("type") == "40ft" for c in container_plan_list)
            or bool(layout_persist.get("suggest_40ft", False)),
            "warnings": v3.get("warnings") or [],
            "container_optimization": v3.get("container_optimization") or {},
        },
    }


@app.delete("/api/projects/{project_id}/crates/")
def delete_all_crates(project_id: int):
    crates = list(crates_col.find({"project_id": project_id}, {"id": 1}))
    crate_ids = [crate["id"] for crate in crates]
    if crate_ids:
        assignments_col.delete_many({"project_id": project_id, "crate_id": {"$in": crate_ids}})
    crates_col.delete_many({"project_id": project_id})
    clear_manual_container_plan(project_id)
    return {"message": "ok"}


@app.delete("/api/crates/{crate_id}")
def delete_crate(crate_id: int):
    crate = crates_col.find_one({"id": crate_id}, {"project_id": 1, "locked": 1})
    if not crate:
        return {"message": "ok"}
    if crate.get("locked"):
        raise HTTPException(status_code=400, detail="Cannot delete a locked crate. Unlock first.")
    assignments_col.delete_many({"project_id": crate["project_id"], "crate_id": crate_id})
    crates_col.delete_one({"id": crate_id})
    clear_manual_container_plan(crate["project_id"])
    return {"message": "ok"}


@app.post("/api/crates/assign")
def assign_piece(data: Dict[str, Any]):
    piece = pieces_col.find_one({"id": int(data["piece_id"])}, {"project_id": 1})
    crate = crates_col.find_one({"id": int(data["crate_id"])}, {"project_id": 1, "locked": 1})
    if not piece or not crate:
        return {"message": "ok"}

    project_id = piece["project_id"]
    if crate.get("locked"):
        raise HTTPException(status_code=400, detail="Target crate is locked — unlock before assigning parts.")
    existing = assignments_col.find_one({"project_id": project_id, "piece_id": piece["id"]})
    if existing:
        src = crates_col.find_one({"project_id": project_id, "id": existing["crate_id"]}, {"locked": 1})
        if src and src.get("locked"):
            raise HTTPException(status_code=400, detail="That part is in a locked crate — unlock before moving.")
    if existing:
        assignments_col.update_one(
            {"id": existing["id"]},
            {"$set": {"crate_id": crate["id"], "assigned_at": utc_now()}},
        )
    else:
        assignments_col.insert_one(
            {
                "id": next_sequence("assignment"),
                "project_id": project_id,
                "piece_id": piece["id"],
                "crate_id": crate["id"],
                "assigned_at": utc_now(),
            }
        )
    clear_manual_container_plan(project_id)
    return {"message": "ok"}


@app.post("/api/crates/unassign")
def unassign_piece(data: Dict[str, Any]):
    piece = pieces_col.find_one({"id": int(data["piece_id"])}, {"project_id": 1})
    if not piece:
        return {"message": "ok"}
    ex = assignments_col.find_one({"project_id": piece["project_id"], "piece_id": piece["id"]})
    if ex:
        src = crates_col.find_one({"project_id": piece["project_id"], "id": ex["crate_id"]}, {"locked": 1})
        if src and src.get("locked"):
            raise HTTPException(status_code=400, detail="Unassign blocked — crate is locked.")
    assignments_col.delete_many({"project_id": piece["project_id"], "piece_id": piece["id"]})
    clear_manual_container_plan(piece["project_id"])
    return {"message": "ok"}


@app.post("/api/projects/{project_id}/crates/assign-family")
def assign_family(project_id: int, data: Dict[str, Any]):
    """
    Atomically reassign all pieces in a family to a target crate.
    If crate_id is None / omitted the family is unassigned.
    """
    piece_ids: List[int] = [int(x) for x in (data.get("piece_ids") or [])]
    crate_db_id = data.get("crate_id")  # DB integer id; None = unassign

    if not piece_ids:
        raise HTTPException(status_code=400, detail="piece_ids required")

    for pid in piece_ids:
        a = assignments_col.find_one({"project_id": project_id, "piece_id": pid})
        if not a:
            continue
        src = crates_col.find_one({"project_id": project_id, "id": a["crate_id"]}, {"locked": 1})
        if src and src.get("locked"):
            raise HTTPException(
                status_code=400,
                detail="A selected part is in a locked crate — unlock before moving the bundle.",
            )

    # Remove existing assignments for these pieces
    assignments_col.delete_many({"project_id": project_id, "piece_id": {"$in": piece_ids}})

    if crate_db_id is not None:
        crate_db_id = int(crate_db_id)
        crate = crates_col.find_one({"id": crate_db_id, "project_id": project_id}, {"_id": 0, "id": 1, "locked": 1})
        if not crate:
            raise HTTPException(status_code=404, detail="Crate not found")
        if crate.get("locked"):
            raise HTTPException(status_code=400, detail="Target crate is locked.")
        for pid in piece_ids:
            assignments_col.insert_one({
                "id": next_sequence("assignment"),
                "project_id": project_id,
                "piece_id": pid,
                "crate_id": crate_db_id,
                "assigned_at": utc_now(),
            })

    clear_manual_container_plan(project_id)
    return {"ok": True, "assigned": len(piece_ids) if crate_db_id is not None else 0, "unassigned": len(piece_ids) if crate_db_id is None else 0}


@app.post("/api/projects/{project_id}/crates/planner-recompute")
def planner_recompute_endpoint(project_id: int):
    """Recompute v3 crate dims/layers/weights and multi-container layout after manual moves."""
    project = projects_col.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    status = project.get("status") or "draft"
    if status not in {"approved_for_packing", "crate_planned", "container_planned"}:
        raise HTTPException(
            status_code=400,
            detail="Approve the project for packing before recalculating the planner.",
        )

    pieces = list(pieces_col.find({"project_id": project_id}, {"_id": 0}))
    crates = list(crates_col.find({"project_id": project_id}, {"_id": 0}))
    assignments = list(assignments_col.find({"project_id": project_id}, {"_id": 0}))
    if not any(c.get("packing_mode") == "v3" for c in crates):
        raise HTTPException(status_code=400, detail="No v3 crates in this project.")

    from .services.planner_v3.recompute_layout import run_planner_recompute

    result = run_planner_recompute(project_id, project, pieces, crates, assignments)
    for cid, patch in result["crate_updates"]:
        crates_col.update_one(
            {"project_id": project_id, "id": cid},
            {"$set": {**patch, "updated_at": utc_now()}},
        )

    projects_col.update_one(
        {"id": project_id},
        {
            "$set": {
                "manual_container_plan": result["manual_plan"],
                "planner_v3_layout": result["layout_persist"],
                "planner_v3_containers": result["enriched_layouts"],
                "planner_v3_summary": result["summary"],
                "planner_v3_container_optimization": result.get("container_optimization") or {},
                "updated_at": utc_now(),
            }
        },
    )
    return result["response"]


@app.get("/api/projects/{project_id}/crates/assignments")
def get_assignments(project_id: int):
    assignments = sorted(
        assignments_col.find({"project_id": project_id}, {"_id": 0}),
        key=lambda doc: doc["id"],
    )
    return [assignment_response(assignment) for assignment in assignments]


@app.get("/api/projects/{project_id}/container-plan")
def get_manual_container_plan(project_id: int):
    project = projects_col.find_one({"id": project_id}, {"_id": 0})
    return (project or {}).get("manual_container_plan") or {"containers": []}


@app.put("/api/projects/{project_id}/container-plan")
def save_manual_container_plan(project_id: int, data: Dict[str, Any]):
    containers = data.get("containers", [])
    sanitized_containers = []
    for index, container in enumerate(containers):
        sanitized_containers.append(
            {
                "id": container.get("id") or f"MANUAL-{index + 1:03d}",
                "type": container.get("type", "40ft"),
                "placements": [
                    {
                        "crate_id": placement.get("crate_id"),
                        "x": float(placement.get("x", 0) or 0),
                        "y": float(placement.get("y", 0) or 0),
                        "rotated": bool(placement.get("rotated", False)),
                        "stack_level": int(placement.get("stack_level", 0) or 0),
                        "loading_order": int(placement.get("loading_order", placement_index + 1) or (placement_index + 1)),
                        "unload_order": int(placement.get("unload_order", placement_index + 1) or (placement_index + 1)),
                    }
                    for placement_index, placement in enumerate(container.get("placements", []))
                    if placement.get("crate_id")
                ],
            }
        )
    projects_col.update_one({"id": project_id}, {"$set": {"manual_container_plan": {"containers": sanitized_containers}, "updated_at": utc_now()}})
    return {"message": "ok"}


@app.delete("/api/projects/{project_id}/container-plan")
def delete_manual_container_plan(project_id: int):
    clear_manual_container_plan(project_id)
    return {"message": "ok"}


@app.get("/api/projects/{project_id}/crates/insights")
def get_crate_insights(project_id: int):
    try:
        return crate_insights(project_id)
    except Exception as e:
        print(f"INSIGHTS ERROR for project {project_id}: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to compute insights: {str(e)}")


# ── Upload PDF → Parse ────────────────────────────────────────────────────────

@app.post("/api/projects/{project_id}/upload-pdf/")
async def upload_pdf(project_id: int, file: UploadFile = File(...)):
    project = projects_col.find_one({"id": project_id}, {"_id": 0}) or {}
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
    pdf_bytes = await file.read()
    if len(pdf_bytes) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 20 MB)")

    result = parse_pdf(pdf_bytes, project)

    # Check for similar existing drawings in the project
    existing_drawings = list({p.get("drawing") or "" for p in pieces_col.find({"project_id": project_id}, {"drawing": 1})})
    extracted_drawing = (result.get("metadata", {}).get("drawing", "")
                         or (result["rows"][0].get("drawing", "") if result.get("rows") else ""))
    similar_drawing = None
    if extracted_drawing and existing_drawings:
        for d in existing_drawings:
            if d and (d.lower() == extracted_drawing.lower() or
                      extracted_drawing.lower().startswith(d.lower()[:4]) or
                      d.lower().startswith(extracted_drawing.lower()[:4])):
                similar_drawing = d
                break

    return {
        **result,
        "file_name": file.filename,
        "similar_drawing": similar_drawing,
        "parser_summary": result.get("metadata", {}).get("page_summary", {}),
    }


# ── Upload Drafts CRUD ────────────────────────────────────────────────────────

class DraftCreate(BaseModel):
    name: str = "Untitled Draft"
    rows: List[Dict] = []
    file_names: List[str] = []


class DraftUpdate(BaseModel):
    name: str = ""
    rows: List[Dict] = []


@app.get("/api/projects/{project_id}/drafts/")
def list_drafts(project_id: int):
    drafts = list(upload_drafts_col.find({"project_id": project_id}, {"_id": 0}))
    return sorted(drafts, key=lambda d: d.get("updated_at") or d.get("created_at") or "", reverse=True)


@app.post("/api/projects/{project_id}/drafts/")
def create_draft(project_id: int, body: DraftCreate):
    draft_id = next_sequence("upload_draft")
    now = utc_now()
    doc = {
        "id": draft_id,
        "project_id": project_id,
        "name": body.name or "Untitled Draft",
        "rows": body.rows,
        "file_names": body.file_names,
        "row_count": len(body.rows),
        "created_at": now,
        "updated_at": now,
    }
    upload_drafts_col.insert_one(doc)
    return {"id": draft_id, "name": doc["name"], "row_count": doc["row_count"], "created_at": now}


@app.get("/api/drafts/{draft_id}")
def get_draft(draft_id: int):
    draft = upload_drafts_col.find_one({"id": draft_id}, {"_id": 0})
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    return draft


@app.put("/api/drafts/{draft_id}")
def update_draft(draft_id: int, body: DraftUpdate):
    now = utc_now()
    update_fields: Dict = {"updated_at": now}
    if body.name:
        update_fields["name"] = body.name
    if body.rows is not None:
        update_fields["rows"] = body.rows
        update_fields["row_count"] = len(body.rows)
    upload_drafts_col.find_one_and_update(
        {"id": draft_id}, {"$set": update_fields}, return_document=ReturnDocument.AFTER
    )
    return {"ok": True, "updated_at": now}


@app.delete("/api/drafts/{draft_id}")
def delete_draft(draft_id: int):
    upload_drafts_col.delete_one({"id": draft_id})
    return {"ok": True}


@app.get("/api/projects/{project_id}/export-source-data")
def export_source_data(project_id: int):
    """Export source data (project info + pieces) as Excel for proforma invoices."""
    project = projects_col.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    pieces = sorted(pieces_col.find({"project_id": project_id}, {"_id": 0}), key=lambda doc: doc["id"])
    if not pieces:
        raise HTTPException(status_code=400, detail="No parts in this project yet.")

    try:
        import openpyxl
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Project Info"
        ws1.append(["StoneDesk — Source Data Export"])
        ws1.append([])
        ws1.append(["Project Name", project.get("name", "")])
        ws1.append(["Customer", project.get("customer", "")])
        ws1.append(["Job Number", project.get("job_number", "")])
        ws1.append(["Material", project.get("material", "Granite")])
        ws1.append(["Thickness", project.get("thickness", "3CM")])
        ws1.append(["Export Date", date.today().isoformat()])
        ws1.append([])
        total_qty = sum(int(p.get("qty", 1) or 1) for p in pieces)
        total_sqft = sum((float(p.get("length", 0)) * float(p.get("width", 0)) / 144.0) * int(p.get("qty", 1) or 1) for p in pieces)
        total_weight = sum(
            planning_piece_weight(p, project.get("material", "Granite"), project.get("thickness", "3CM"), project.get("stone_color", "") or "") * int(p.get("qty", 1) or 1)
            for p in pieces
        )
        ws1.append(["Total Parts", total_qty])
        ws1.append(["Total Sq Ft", round(total_sqft, 2)])
        ws1.append(["Total Weight (kg)", round(total_weight, 2)])

        ws2 = wb.create_sheet("Parts List")
        ws2.append([
            # Identity
            "Part #", "Description", "Category", "Drawing", "Unit",
            # Location
            "Building", "Floor", "Flat",
            # Dimensions
            "Length (in)", "Width (in)", "Qty", "Sq Ft", "Weight (kg)",
            # Sink
            "Sink Type", "Sink Cutouts", "Tap Holes", "Grooves",
            # Edge
            "Edge Type", "Edge Sides", "Edge Polish (in)", "Edge Per-Side", "Edge Manual Note",
            # Radius
            "Radius (in)", "Radius Corners", "No. Corners",
            # Shape & logistics
            "Shape Type", "Fragility", "Orientation", "Delivery Priority", "Stack Preference",
            "Weight Override (kg)", "Notes",
        ])
        _bold_row(ws2, 1)
        mat = project.get("material", "Granite")
        thick = project.get("thickness", "3CM")
        _color = project.get("stone_color", "") or ""
        for p in pieces:
            qty = int(p.get("qty", 1) or 1)
            sqft = round((float(p.get("length", 0)) * float(p.get("width", 0)) / 144.0) * qty, 2)
            wt = round(planning_piece_weight(p, mat, thick, _color) * qty, 2)
            rc = p.get("radius_corners") or {}
            active_corners = sum(1 for v in rc.values() if v)
            ws2.append([
                p.get("part_no", ""), p.get("part", ""), p.get("category", ""), p.get("drawing", ""), p.get("unit", ""),
                p.get("building", ""), p.get("floor", ""), p.get("flat", ""),
                p.get("length", 0), p.get("width", 0), qty, sqft, wt,
                p.get("sink_type", "No Sink"), p.get("sink_cut", "-"), p.get("tap_holes", "-"), p.get("grooves", "-"),
                p.get("edge", "None"), p.get("edge_area", ""), round(float(p.get("edge_polish_machine", 0) or 0), 2),
                _fmt_edge_map(p.get("edge_map")), p.get("edge_polish_manual", ""),
                p.get("radius_value", ""), _fmt_radius_corners(rc), active_corners or "",
                p.get("shape_type", ""),
                p.get("fragility", "Standard"), p.get("orientation", "Auto"),
                p.get("delivery_priority", "Standard"), p.get("stack_preference", "Auto"),
                p.get("weight_override", 0) or "",
                p.get("notes", ""),
            ])

        _hide_empty_columns(ws2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=SourceData_{project.get('name', 'Project')}_{date.today()}.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"SOURCE DATA EXPORT ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Source data export failed: {str(e)}")


@app.get("/api/projects/{project_id}/export")
def export_excel(project_id: int):
    project = projects_col.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    pieces = sorted(pieces_col.find({"project_id": project_id}, {"_id": 0}), key=lambda doc: doc["id"])
    if not pieces:
        raise HTTPException(status_code=400, detail="Cannot export: no parts in this project. Add parts first.")

    crates_raw = list(crates_col.find({"project_id": project_id}, {"_id": 0}))
    if not crates_raw:
        raise HTTPException(status_code=400, detail="Cannot export: no crates generated. Generate a plan first.")

    try:
        assignments = list(assignments_col.find({"project_id": project_id}, {"_id": 0}))
        pieces_by_crate = pieces_grouped_by_crate(pieces, assignments)
        snapshot = project_planning_snapshot(project_id)
        loading_plan = build_container_plan(
            snapshot["crate_rows"],
            manual_plan=project.get("manual_container_plan"),
            preferred_mode=project.get("preferred_container_mode"),
        )

        import openpyxl
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["StoneDesk Planning Summary"])
        ws1.append([])
        ws1.append(["Project", project.get("name", "")])
        ws1.append(["Customer", project.get("customer", "")])
        ws1.append(["Job Number", project.get("job_number", "")])
        ws1.append(["Material", project.get("material", "Granite")])
        ws1.append(["Thickness", project.get("thickness", "3CM")])
        ws1.append(["Crate Wood Type", project.get("crate_wood_type", "Pine")])
        ws1.append(["Crate Wood Thickness (in)", project.get("crate_wood_thickness", 1.25)])
        ws1.append(["Export Date", date.today().isoformat()])
        ws1.append([])
        ws1.append(["KPI", "Value"])
        ws1.append(["Crates Created", len(snapshot["crate_rows"])])
        ws1.append(["Total Stone Weight (kg)", snapshot.get("total_weight", 0)])
        ws1.append(["Shipment Weight incl. Tare (kg)", snapshot.get("total_gross_weight", 0)])
        ws1.append(["Total Sq Ft", snapshot.get("total_sqft", 0)])
        ws1.append(["Average Fill %", snapshot.get("average_fill_percent", 0)])
        ws1.append(["Average Gross Weight Util %", snapshot.get("average_gross_utilization", 0)])

        recommendation = loading_plan.get("recommendation") or {}
        ws1.append(["Recommended Booking", recommendation.get("booking_action", "N/A")])
        ws1.append(["Recommended Mode", recommendation.get("mode_label", "N/A")])
        ws1.append(["Recommendation Reason", recommendation.get("reason", "")])
        ws1.append([])
        ws1.append(["Alternative Options"])
        for option in loading_plan.get("options", []):
            option_line = (
                f"{option.get('label', '?')} | feasible={option.get('feasible', False)} | "
                f"avg weight util {option.get('average_weight_utilization', 0):.0f}% | "
                f"avg length util {option.get('average_length_utilization', 0):.0f}% | "
                f"cost index {option.get('cost_index', 0):.2f}"
            )
            ws1.append([option_line])

        ws2 = wb.create_sheet("Crates")
        ws2.append(
            [
                "Crate ID",
                "Crate Name",
                "Destination Group",
                "Family Group",
                "Locked",
                "Custom",
                "Int L",
                "Int W",
                "Int H",
                "Ext L",
                "Ext W",
                "Ext H",
                "Wood Type",
                "Wood Thickness",
                "Tare Wt",
                "Net Wt",
                "Gross Wt",
                "Fill %",
                "Gross Util %",
                "Center of Gravity",
                "Forklift Entry",
                "Stackable",
                "Reinforcement",
                "Reserved Space %",
                "Status",
                "Handling Notes",
            ]
        )
        for crate in snapshot["crate_rows"]:
            cog = crate.get("center_of_gravity") or {"x": 0, "y": 0, "z": 0}
            ws2.append(
                [
                    crate.get("crate_id", ""),
                    crate.get("name", ""),
                    crate.get("destination_group", ""),
                    crate.get("family_group", ""),
                    crate.get("locked", False),
                    crate.get("custom", False),
                    crate.get("internal_length", 0),
                    crate.get("internal_width", 0),
                    crate.get("internal_height", 0),
                    crate.get("external_length", 0),
                    crate.get("external_width", 0),
                    crate.get("external_height", 0),
                    crate.get("wood_type", ""),
                    crate.get("wood_thickness", 0),
                    crate.get("tare_weight", 0),
                    crate.get("total_weight", 0),
                    crate.get("gross_weight", 0),
                    crate.get("fill_percent", 0),
                    crate.get("gross_utilization", 0),
                    f"{cog.get('x', 0)}, {cog.get('y', 0)}, {cog.get('z', 0)}",
                    crate.get("forklift_entry", ""),
                    crate.get("stackable", False),
                    crate.get("reinforcement", False),
                    crate.get("reserved_space_pct", 0),
                    crate.get("efficiency_status", ""),
                    crate.get("handling_notes", ""),
                ]
            )

        ws3 = wb.create_sheet("Crate Contents")
        ws3.append([
            # Crate
            "Crate ID", "Crate Name",
            # Identity
            "Piece ID", "Part #", "Description", "Category", "Drawing", "Unit",
            # Location
            "Building", "Floor", "Flat",
            # Dimensions
            "Length (in)", "Width (in)", "Qty", "Stone Wt (kg)",
            # Sink
            "Sink Type", "Cutouts", "Tap Holes", "Grooves",
            # Edge
            "Edge Type", "Edge Sides", "Edge Polish (in)", "Edge Per-Side", "Edge Manual Note",
            # Radius
            "Radius (in)", "Radius Corners", "No. Corners",
            # Shape & logistics
            "Shape Type", "Fragility", "Orientation", "Delivery Priority", "Stack Preference",
            "Weight Override (kg)", "Notes",
        ])
        _bold_row(ws3, 1)
        _mat = project.get("material", "Granite")
        _thick = project.get("thickness", "3CM")
        _color2 = project.get("stone_color", "") or ""
        for crate in snapshot["crate_rows"]:
            crate_pieces = pieces_by_crate.get(crate["id"], [])
            for piece in crate_pieces:
                rc = piece.get("radius_corners") or {}
                active_corners = sum(1 for v in rc.values() if v)
                ws3.append([
                    crate.get("crate_id", ""), crate.get("name", ""),
                    piece.get("id", ""), piece.get("part_no", ""), piece.get("part", ""),
                    piece.get("category", ""), piece.get("drawing", ""), piece.get("unit", ""),
                    piece.get("building", ""), piece.get("floor", ""), piece.get("flat", ""),
                    piece.get("length", 0), piece.get("width", 0), piece.get("qty", 1),
                    round(planning_piece_weight(piece, _mat, _thick, _color2), 2),
                    piece.get("sink_type", "No Sink"), piece.get("sink_cut", "-"),
                    piece.get("tap_holes", "-"), piece.get("grooves", "-"),
                    piece.get("edge", "None"), piece.get("edge_area", ""),
                    round(float(piece.get("edge_polish_machine", 0) or 0), 2),
                    _fmt_edge_map(piece.get("edge_map")), piece.get("edge_polish_manual", ""),
                    piece.get("radius_value", ""), _fmt_radius_corners(rc), active_corners or "",
                    piece.get("shape_type", ""),
                    piece.get("fragility", "Standard"), piece.get("orientation", "Auto"),
                    piece.get("delivery_priority", "Standard"), piece.get("stack_preference", "Auto"),
                    piece.get("weight_override", 0) or "",
                    piece.get("notes", ""),
                ])

        _hide_empty_columns(ws3)

        ws4 = wb.create_sheet("Container Plan")
        ws4.append(
            [
                "Container ID",
                "Type",
                "Crate ID",
                "Crate Name",
                "Destination",
                "X",
                "Y",
                "Length",
                "Width",
                "Gross Wt",
                "Loading Order",
                "Unload Order",
                "Rotated",
                "Container Warnings",
            ]
        )
        for container in loading_plan.get("containers", []):
            warnings = "; ".join(container.get("warnings", []))
            for placement in container.get("placements", []):
                ws4.append(
                    [
                        container.get("id", ""),
                        container.get("type", ""),
                        placement.get("crate_id", ""),
                        placement.get("name", ""),
                        placement.get("destination_group", ""),
                        placement.get("x", 0),
                        placement.get("y", 0),
                        placement.get("length", 0),
                        placement.get("width", 0),
                        placement.get("weight", 0),
                        placement.get("loading_order", 0),
                        placement.get("unload_order", 0),
                        placement.get("rotated", False),
                        warnings,
                    ]
                )

        ws5 = wb.create_sheet("Exceptions - Warnings")
        ws5.append(["Scope", "ID", "Name", "Severity", "Message"])
        for row in snapshot.get("exception_rows", []):
            ws5.append([row.get("scope", ""), row.get("id", ""), row.get("name", ""), row.get("severity", ""), row.get("message", "")])
        for row in snapshot.get("underfilled_crates", []):
            ws5.append(["crate", row.get("crate_id", ""), row.get("name", ""), row.get("status", ""), row.get("suggestion", "")])
        for container in loading_plan.get("containers", []):
            for warning in container.get("warnings", []):
                ws5.append(["container", container.get("id", ""), container.get("type", ""), "yellow", warning])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=StoneDesk_{date.today()}.xlsx"},
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"EXPORT ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
