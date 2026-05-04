from openpyxl import Workbook
from io import BytesIO
from typing import List
from ..models import Piece, Crate, Assignment

def create_excel_export(pieces: List[Piece], crates: List[Crate], assignments: List[Assignment],
                        project_name: str, customer: str, job_number: str, material: str,
                        thickness: str, export_date: str) -> BytesIO:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Line Items"
    headers = ["#", "Part", "Category", "Material", "Thickness", "Drawing", "Unit", "Length", "Depth", "Qty", "Sink", "Cutouts", "Tap", "Grooves", "Edge", "Radius", "SqFt ea", "SqFt tot", "kg ea", "kg tot", "Building", "Floor", "Flat", "Notes"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    for idx, p in enumerate(pieces, 1):
        sqft = (p.length * p.width) / 144
        ws1.cell(row=idx+1, column=1, value=idx)
        ws1.cell(row=idx+1, column=2, value=p.part)
        ws1.cell(row=idx+1, column=3, value=p.category)
        ws1.cell(row=idx+1, column=4, value=material)
        ws1.cell(row=idx+1, column=5, value=thickness)
        ws1.cell(row=idx+1, column=6, value=p.drawing or "")
        ws1.cell(row=idx+1, column=7, value=p.unit or "")
        ws1.cell(row=idx+1, column=8, value=p.length)
        ws1.cell(row=idx+1, column=9, value=p.width)
        ws1.cell(row=idx+1, column=10, value=p.qty)
        ws1.cell(row=idx+1, column=11, value=p.sink_type)
        ws1.cell(row=idx+1, column=12, value=p.sink_cut or "-")
        ws1.cell(row=idx+1, column=13, value=p.tap_holes or "-")
        ws1.cell(row=idx+1, column=14, value=p.grooves or "-")
        ws1.cell(row=idx+1, column=15, value=p.edge or "-")
        ws1.cell(row=idx+1, column=16, value=p.radius or "-")
        ws1.cell(row=idx+1, column=17, value=round(sqft, 2))
        ws1.cell(row=idx+1, column=18, value=round(sqft * p.qty, 2))
        ws1.cell(row=idx+1, column=19, value=round(sqft * 7.5, 2))
        ws1.cell(row=idx+1, column=20, value=round(sqft * 7.5 * p.qty, 2))
        ws1.cell(row=idx+1, column=21, value=p.building or "")
        ws1.cell(row=idx+1, column=22, value=p.floor or "")
        ws1.cell(row=idx+1, column=23, value=p.flat or "")
        ws1.cell(row=idx+1, column=24, value=p.notes or "")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output